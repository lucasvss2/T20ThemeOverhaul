"""
Exporta a planilha "T20 - Tabela de geração de tesouros.xlsx" para um arquivo
de dados TS (src/treasure/treasure-data.ts), decodificando as células que o
Excel converteu em datas (frações de ND e faixas de d% tipo "01-10").

Uso: python scripts/export-treasure.py
"""
import openpyxl, datetime, json, sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SRC = r"E:\Foundry\T20\T20 - Tabela de geração de tesouros.xlsx"
OUT = r"E:\Foundry\New Lib\src\treasure\treasure-data.ts"

wb = openpyxl.load_workbook(SRC, data_only=True)


def is_dt(v):
    return isinstance(v, datetime.datetime)


def s(v):
    if v is None:
        return ""
    if is_dt(v):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def decode_range(v):
    """Retorna [min,max] (1..100) ou None para vazio/—. Decodifica datas mangladas."""
    if v is None:
        return None
    if is_dt(v):
        return [v.day, v.month]
    t = str(v).strip()
    if t in ("", "—", "-", "–"):
        return None
    if isinstance(v, float) and v.is_integer():
        t = str(int(v))
    if t.endswith(".0"):
        t = t[:-2]
    try:
        if "-" in t:
            a, b = t.split("-", 1)
            a, b = a.strip(), b.strip()
            if a.isdigit() and b.isdigit():
                return [int(a), int(b)]
            return None
        if t.isdigit():
            return [int(t), int(t)]
    except ValueError:
        return None
    return None


def decode_nd(v):
    if is_dt(v):
        return f"{v.day}/{v.month}"
    t = s(v)
    if t.endswith(".0"):
        t = t[:-2]
    return t


def clean(v):
    """Texto de célula 'normal' (nome, livro, valor…) — datas viram '' (não esperadas aqui)."""
    if v is None:
        return ""
    if is_dt(v):
        return ""  # não deveria ocorrer em colunas de texto
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# ── Main table: Tesouro por ND ────────────────────────────────────────────────
def parse_main():
    ws = wb["Tesouro por ND"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    cur = None
    for r in rows[1:]:  # skip header
        nd_cell = r[0]
        # nova ND quando col A preenchida
        if nd_cell is not None and s(nd_cell) != "":
            nd = decode_nd(nd_cell)
            # parar no rodapé (+%/2D notes começam sem dinheiro/itens)
            if nd.startswith("+") or nd.startswith("2D"):
                cur = None
                continue
            cur = {"nd": nd, "dinheiro": [], "itens": []}
            out.append(cur)
        if cur is None:
            continue
        dr = decode_range(r[1])
        if dr:
            cur["dinheiro"].append({"range": dr, "result": clean(r[2])})
        ir = decode_range(r[4])
        if ir:
            cur["itens"].append({"range": ir, "result": clean(r[5])})
    return out


# ── Tabela simples d% → colunas ───────────────────────────────────────────────
def parse_simple(sheet, col_range, col_fields, header_rows=1):
    """col_range = índice (0-based) da coluna de d%. col_fields = lista de (chave, idx)."""
    ws = wb[sheet]
    out = []
    for r in list(ws.iter_rows(values_only=True))[header_rows:]:
        rng = decode_range(r[col_range])
        if not rng:
            continue
        entry = {"range": rng}
        for key, idx in col_fields:
            entry[key] = clean(r[idx]) if idx < len(r) else ""
        out.append(entry)
    return out


# ── Tabela de 3 blocos (Equipamentos/Superiores/Mágicos/Acessórios) ───────────
def parse_blocks(sheet, blocks, header_rows=2):
    """blocks = lista de (chave_bloco, col_inicial_0based, [(chave,offset)...])."""
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))[header_rows:]
    out = {}
    for bkey, base, fields in blocks:
        lst = []
        for r in rows:
            rng = decode_range(r[base]) if base < len(r) else None
            if not rng:
                continue
            entry = {"range": rng}
            for key, off in fields:
                idx = base + off
                entry[key] = clean(r[idx]) if idx < len(r) else ""
            lst.append(entry)
        out[bkey] = lst
    return out


# ── Riquezas: escada de valor com 3 mapeamentos de d% (menor/média/maior) ─────
def parse_riquezas():
    ws = wb["Riquezas"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    # header row1 = título, row2 = "Menor|Média|Maior|Valor|Exemplos"
    for r in rows[2:]:
        valor = clean(r[3]) if len(r) > 3 else ""
        menor = decode_range(r[0])
        media = decode_range(r[1])
        maior = decode_range(r[2])
        exemplos = (str(r[4]).strip() if len(r) > 4 and r[4] is not None else "")
        if not (menor or media or maior) and not valor:
            continue
        out.append({"valor": valor, "exemplos": exemplos,
                    "menor": menor, "media": media, "maior": maior})
    return out


data = {
    "main": parse_main(),
    "itensDiversos": parse_simple("Itens Diversos", 0, [("item", 1), ("livro", 2), ("pagina", 3)]),
    "pocoes": parse_simple("Poções", 0, [("nome", 1), ("preco", 2), ("livro", 3), ("pagina", 4)]),
    "equipamentos": parse_blocks("Equipamentos", [
        ("arma", 0, [("nome", 1), ("livro", 2), ("pagina", 3)]),
        ("armadura", 5, [("nome", 1), ("livro", 2), ("pagina", 3)]),
        ("esoterico", 10, [("nome", 1), ("livro", 2), ("pagina", 3)]),
    ]),
    "superiores": parse_blocks("Superiores", [
        ("arma", 0, [("nome", 1), ("livro", 2), ("pagina", 3)]),
        ("armadura", 5, [("nome", 1), ("livro", 2), ("pagina", 3)]),
        ("esoterico", 10, [("nome", 1), ("livro", 2), ("pagina", 3)]),
    ]),
    "magicos": parse_blocks("Mágicos", [
        ("arma", 0, [("nome", 1), ("livro", 2), ("pagina", 3)]),
        ("armadura", 5, [("nome", 1), ("livro", 2), ("pagina", 3)]),
        ("esoterico", 10, [("nome", 1), ("livro", 2), ("pagina", 3)]),
    ]),
    "acessorios": parse_blocks("Mágicos (Acessórios)", [
        ("menor", 0, [("nome", 1), ("preco", 2), ("livro", 3), ("pagina", 4)]),
        ("medio", 6, [("nome", 1), ("preco", 2), ("livro", 3), ("pagina", 4)]),
        ("maior", 12, [("nome", 1), ("preco", 2), ("livro", 3), ("pagina", 4)]),
    ]),
    "riquezas": parse_riquezas(),
}


# ── Validação de cobertura d% ─────────────────────────────────────────────────
def coverage(name, rows, key="range"):
    rngs = [e[key] for e in rows if e.get(key)]
    if not rngs:
        return f"  [{name}] VAZIO"
    problems = []
    inv = [r for r in rngs if r[0] > r[1]]
    if inv:
        problems.append(f"invertidas={inv[:4]}")
    cov = set()
    for a, b in rngs:
        for x in range(min(a, b), max(a, b) + 1):
            cov.add(x)
    miss = [x for x in range(1, 101) if x not in cov]
    if miss:
        problems.append(f"faltando 1-100={miss[:10]}")
    return f"  [{name}] {len(rngs)} faixas " + ("OK" if not problems else "PROBLEMAS: " + "; ".join(problems))


print("=== VALIDAÇÃO DE COBERTURA ===")
for env in data["main"]:
    print(coverage(f"ND {env['nd']} dinheiro", env["dinheiro"]))
    print(coverage(f"ND {env['nd']} itens", env["itens"]))
for k in ["itensDiversos", "pocoes"]:
    print(coverage(k, data[k]))
for k in ["equipamentos", "superiores", "magicos", "acessorios"]:
    for bk, lst in data[k].items():
        print(coverage(f"{k}.{bk}", lst))
for cat in ["menor", "media", "maior"]:
    print(coverage(f"riquezas.{cat}", [{"range": r[cat]} for r in data["riquezas"] if r[cat]]))

print("\n=== TAMANHOS ===")
print("main NDs:", len(data["main"]))
for k in ["itensDiversos", "pocoes", "riquezas"]:
    print(k, len(data[k]))
for k in ["equipamentos", "superiores", "magicos", "acessorios"]:
    print(k, {bk: len(v) for bk, v in data[k].items()})

# ── Emite TS ──────────────────────────────────────────────────────────────────
HEADER = '''/**
 * Dados da tabela de geração de tesouros (Tormenta20).
 *
 * GERADO AUTOMATICAMENTE por scripts/export-treasure.py a partir de
 * "T20 - Tabela de geração de tesouros.xlsx". NÃO EDITAR À MÃO — para
 * atualizar/expandir, edite a planilha e rode o script novamente.
 *
 * Faixas d% são [min, max] inclusivas. Crédito da planilha: Guilherme Dei
 * Svaldi; riquezas adicionais: Rafael Dei Svaldi.
 */

export interface DPRow { range: [number, number]; result: string }
export interface NDEntry { nd: string; dinheiro: DPRow[]; itens: DPRow[] }
export interface ItemRow { range: [number, number]; nome?: string; item?: string; preco?: string; livro?: string; pagina?: string }
export interface RiquezaRow { valor: string; exemplos: string; menor: [number, number] | null; media: [number, number] | null; maior: [number, number] | null }
export interface TreasureData {
    main: NDEntry[];
    itensDiversos: ItemRow[];
    pocoes: ItemRow[];
    equipamentos: Record<string, ItemRow[]>;
    superiores: Record<string, ItemRow[]>;
    magicos: Record<string, ItemRow[]>;
    acessorios: Record<string, ItemRow[]>;
    riquezas: RiquezaRow[];
}

export const TREASURE: TreasureData = '''

with open(OUT, "w", encoding="utf-8") as f:
    f.write(HEADER)
    f.write(json.dumps(data, ensure_ascii=False, indent=1))
    f.write(";\n")

print("\nGerado:", OUT)
